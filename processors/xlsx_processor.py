from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from core.file_utils import make_masked_output_path
from core.models import DetectionItem
from core.pii_detector import PiiDetector
from processors.base_processor import BaseProcessor


class XlsxProcessor(BaseProcessor):
    file_type = "xlsx"
    supported_suffixes = {".xlsx"}

    def __init__(self, detector: PiiDetector | None = None) -> None:
        self.detector = detector or PiiDetector()

    def detect(self, file_path: str | Path) -> list[DetectionItem]:
        input_path = Path(file_path)
        workbook = load_workbook(input_path, data_only=False)
        detections: list[DetectionItem] = []

        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        if isinstance(cell.value, str):
                            source = {
                                "file_path": str(input_path),
                                "file_type": self.file_type,
                                "page_or_sheet": worksheet.title,
                                "location": cell.coordinate,
                            }
                            formula_items = self.detector.detect_text(cell.value, source)
                            for item in formula_items:
                                item.selected = False
                                item.review_status = "제외"
                                item.note = "수식 셀은 자동 마스킹 제외"
                            detections.extend(formula_items)
                        continue
                    if not isinstance(cell.value, str):
                        continue

                    source = {
                        "file_path": str(input_path),
                        "file_type": self.file_type,
                        "page_or_sheet": worksheet.title,
                        "location": cell.coordinate,
                    }
                    detections.extend(self.detector.detect_text(cell.value, source))

        return detections

    def apply_masking(self, file_path: str | Path, output_dir: str | Path, items: list[DetectionItem]) -> str:
        input_path = Path(file_path)
        output_folder = Path(output_dir)
        output_path = make_masked_output_path(input_path, output_folder)
        workbook = load_workbook(input_path, data_only=False)

        grouped_items = self._group_selected_items(items, input_path)
        for (sheet_name, cell_address), selected_items in grouped_items.items():
            if sheet_name not in workbook.sheetnames:
                continue
            cell = workbook[sheet_name][cell_address]
            if cell.data_type == "f" or not isinstance(cell.value, str):
                continue
            cell.value = self._replace_selected(cell.value, selected_items)

        output_folder.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return str(output_path)

    def save_masked_copy(self, input_path: Path, output_path: Path, detections: list[DetectionItem]) -> None:
        workbook = load_workbook(input_path, data_only=False)
        grouped_items = self._group_selected_items(detections, input_path)
        for (sheet_name, cell_address), selected_items in grouped_items.items():
            if sheet_name not in workbook.sheetnames:
                continue
            cell = workbook[sheet_name][cell_address]
            if cell.data_type == "f" or not isinstance(cell.value, str):
                continue
            cell.value = self._replace_selected(cell.value, selected_items)
        workbook.save(output_path)

    @staticmethod
    def _replace_selected(text: str, selected_items: list[DetectionItem]) -> str:
        result = text
        for item in selected_items:
            result = result.replace(item.original_text, item.masked_text, 1)
        return result

    @staticmethod
    def _group_selected_items(items: list[DetectionItem], input_path: Path) -> dict[tuple[str, str], list[DetectionItem]]:
        grouped: dict[tuple[str, str], list[DetectionItem]] = defaultdict(list)
        for item in items:
            if not item.selected:
                continue
            if Path(item.file_path) != input_path:
                continue
            grouped[(item.page_or_sheet, item.location)].append(item)
        return grouped
