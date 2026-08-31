from pathlib import Path

from openpyxl import Workbook, load_workbook

from processors.xlsx_processor import XlsxProcessor


def create_sample_xlsx(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = "홍길동"
    worksheet["A2"] = "010-1234-5678"
    worksheet["A3"] = "teacher@school.kr"
    worksheet["A4"] = "900101-1234567"
    worksheet["A5"] = "서울특별시 강남구 역삼동"
    worksheet["A6"] = "니코초등학교"
    worksheet["A7"] = "홍길동 / 010-1234-5678 / teacher@school.kr"
    workbook.save(path)


def test_xlsx_detects_pii_with_sheet_and_cell_location(tmp_path: Path) -> None:
    input_path = tmp_path / "sample.xlsx"
    create_sample_xlsx(input_path)

    items = XlsxProcessor().detect(input_path)
    types = {item.pii_type for item in items}

    assert {"이름", "전화번호", "이메일", "주민등록번호", "주소", "학교명"}.issubset(types)
    assert any(item.page_or_sheet == "Sheet1" and item.location == "A2" for item in items)

    by_type = {item.pii_type: item for item in items}
    assert by_type["전화번호"].review_status == "자동선택"
    assert by_type["이메일"].review_status == "자동선택"
    assert by_type["이름"].review_status == "확인 필요"
    assert by_type["주소"].review_status == "확인 필요"
    assert by_type["학교명"].review_status == "확인 필요"


def test_xlsx_apply_masking_only_selected_items_and_keeps_original(tmp_path: Path) -> None:
    input_path = tmp_path / "sample.xlsx"
    output_dir = tmp_path / "out"
    create_sample_xlsx(input_path)

    processor = XlsxProcessor()
    items = processor.detect(input_path)
    for item in items:
        item.selected = item.pii_type == "전화번호"

    output_path = Path(processor.apply_masking(input_path, output_dir, items))

    original = load_workbook(input_path).active
    result = load_workbook(output_path).active

    assert output_path.exists()
    assert output_path.name == "masked_sample.xlsx"
    assert original["A2"].value == "010-1234-5678"
    assert result["A2"].value == "010-****-5678"
    assert result["A3"].value == "teacher@school.kr"
    assert result["A7"].value == "홍길동 / 010-****-5678 / teacher@school.kr"


def test_xlsx_uses_user_edited_masked_text(tmp_path: Path) -> None:
    input_path = tmp_path / "sample.xlsx"
    output_dir = tmp_path / "out"
    create_sample_xlsx(input_path)

    processor = XlsxProcessor()
    items = processor.detect(input_path)
    for item in items:
        item.selected = item.original_text == "teacher@school.kr"
        if item.selected:
            item.masked_text = "직접수정@example.com"

    output_path = Path(processor.apply_masking(input_path, output_dir, items))
    result = load_workbook(output_path).active

    assert result["A3"].value == "직접수정@example.com"
